import requests
import re
import os
import logging
from urllib.robotparser import RobotFileParser
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

ROBOTS_DIR = 'robots_files'
REQUEST_HEADERS = {'User-Agent': 'robots-txt-tracker/1.0'}

def clean_url(url):
    """Cleans a URL to be used as a filename."""
    # Remove protocol and www.
    cleaned = re.sub(r'https?://(www\.)?', '', url)
    # Remove paths, queries, etc.
    cleaned = cleaned.split('/')[0]
    return cleaned

def fetch_and_save_robots_txt(urls):
    """Reads URLs from config.txt, fetches robots.txt, and saves them."""
    os.makedirs(ROBOTS_DIR, exist_ok=True)
    for url in urls:
        try:
            # Fetched content is parsed and published to the spreadsheet, so a
            # cleartext fetch is tamperable in transit. All of config.txt is
            # https:// today; warn rather than fail so a site that genuinely
            # has no TLS can still be tracked deliberately.
            if not url.lower().startswith('https://'):
                logger.warning(
                    "Fetching %s over an insecure connection -- response can be "
                    "tampered with in transit; prefer https:// in config.txt", url
                )

            robots_url = f"{url.rstrip('/')}/robots.txt"
            # NOTE: requests.get follows redirects by default. A site could
            # redirect /robots.txt somewhere unexpected. No action taken as
            # this is standard behaviour and generally desirable.
            response = requests.get(robots_url, headers=REQUEST_HEADERS, timeout=10)
            response.raise_for_status()

            filename_base = clean_url(url)
            filename = os.path.join(ROBOTS_DIR, f"{filename_base}.robots.txt")

            with open(filename, 'w', encoding='utf-8') as out_file:
                out_file.write(response.text)
            logger.info("Successfully fetched and saved %s", filename)

        except requests.exceptions.RequestException as e:
            logger.error("Failed to fetch robots.txt from %s: %s", url, e)

def get_user_agents_from_files(robot_files):
    """Extracts all User-Agents from a list of robots.txt files."""
    user_agents = set(['*'])  # Always include the wildcard
    for file_path in robot_files:
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    if line.strip().lower().startswith('user-agent:'):
                        agent = line.split(':', 1)[1].strip().lower()
                        agent = agent.split('#', 1)[0].strip()         # remove comments from name
                        # NOTE: The disallow split is a workaround for malformed robots.txt
                        # files that put Disallow on the same line as User-agent. This could
                        # theoretically strip a legitimate UA name containing "disallow:" but
                        # that is extremely unlikely in practice.
                        agent = agent.split('disallow:', 1)[0].strip() # remove Disallow commands
                        if agent:
                            user_agents.add(agent)
        except FileNotFoundError:
            logger.warning("Could not find %s to extract user agents.", file_path)
    return sorted(list(user_agents))

def detect_directives(file_path):
    """Checks for the presence of Sitemap, License (RSL), and Content-signal directives."""
    found = {'CS Support': False, 'RSL Support': False, 'Sitemap': False}
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                stripped = line.strip().lower()
                if stripped.startswith('content-signal:'):
                    found['CS Support'] = True
                elif stripped.startswith('license:'):
                    found['RSL Support'] = True
                elif stripped.startswith('sitemap:'):
                    found['Sitemap'] = True
    except FileNotFoundError:
        pass
    return found

def update_spreadsheet(urls):
    """Creates or updates a spreadsheet with an analysis of robots.txt files."""
    now = datetime.now()
    spreadsheet_name = f'robots-analysis-{now.year}.xlsx'
    sheet_name = now.strftime('%Y-%m-%d')

    domains = [clean_url(u) for u in urls]
    domain_to_url = dict(zip(domains, urls))

    # Check for collisions where multiple URLs clean to the same domain
    if len(domains) != len(set(domains)):
        seen = set()
        for d, u in zip(domains, urls):
            if d in seen:
                logger.warning("Domain collision: '%s' (from %s) duplicates an earlier entry", d, u)
            seen.add(d)

    # Build file path lookup once for all domains
    robot_file_paths = {d: os.path.join(ROBOTS_DIR, f"{d}.robots.txt") for d in domains}

    if not any(os.path.exists(p) for p in robot_file_paths.values()):
        logger.warning("No robots.txt files found to analyze.")
        return

    user_agents = get_user_agents_from_files(list(robot_file_paths.values()))

    # Detect directives once per domain and cache the results
    directive_cache = {}
    for domain in domains:
        directive_cache[domain] = detect_directives(robot_file_paths[domain])

    # Build directive-presence rows (CS Support, RSL Support, Sitemap)
    directive_labels = ['CS Support', 'RSL Support', 'Sitemap']
    data = []
    for label in directive_labels:
        row = {'User-Agent': label}
        for domain in domains:
            row[domain] = 1 if directive_cache[domain][label] else 0
        data.append(row)

    # Parse each domain's robots.txt once and cache the parser
    parser_cache = {}
    for domain in domains:
        path = robot_file_paths[domain]
        if not os.path.exists(path):
            parser_cache[domain] = None
            continue
        rp = RobotFileParser()
        rp.set_url(f"{domain_to_url[domain].rstrip('/')}/robots.txt")
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            rp.parse(f.readlines())
        parser_cache[domain] = rp

    for ua in user_agents:
        row = {'User-Agent': ua}
        for domain in domains:
            rp = parser_cache[domain]
            if rp is None:
                row[domain] = ''
                continue
            row[domain] = 1 if rp.can_fetch(ua, '/') else 0
        data.append(row)

    if not data:
        logger.warning("No data to write to spreadsheet.")
        return

    df = pd.DataFrame(data)
    df = df.rename(columns={'User-Agent': 'User-Agent / Feature'})
    df = df.set_index('User-Agent / Feature')

    mode = 'a' if os.path.exists(spreadsheet_name) else 'w'
    with pd.ExcelWriter(
        spreadsheet_name,
        engine='openpyxl',
        mode=mode,
        if_sheet_exists='replace'
    ) as writer:
        df.to_excel(writer, sheet_name=sheet_name)

    # Re-open with openpyxl to apply styles and move new sheet to front
    book = load_workbook(spreadsheet_name)
    ws = book[sheet_name]
    book.move_sheet(ws, offset=-len(book.sheetnames) + 1)

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    bold_font = Font(bold=True)

    # Bold the header row
    for cell in ws[1]:
        cell.font = bold_font

    # Bold the 3 feature name cells (rows 2-4, column A)
    for row_idx in range(2, 2 + len(directive_labels)):
        ws.cell(row=row_idx, column=1).font = bold_font

    for row in ws.iter_rows(min_row=2, min_col=2):  # Skip header and index column
        for cell in row:
            if cell.value == 1:
                cell.fill = green_fill
            elif cell.value == 0:
                cell.fill = red_fill

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    book.save(spreadsheet_name)
    logger.info("Spreadsheet '%s' updated with new sheet '%s'.", spreadsheet_name, sheet_name)

def main():
    """Main function to run the script."""
    config_file = 'config.txt'
    if not os.path.exists(config_file):
        logger.error("%s not found.", config_file)
        return

    with open(config_file, 'r') as f:
        urls = [line.strip() for line in f if line.strip()]

    fetch_and_save_robots_txt(urls)
    update_spreadsheet(urls)

if __name__ == "__main__":
    main()
